import os
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

# ---- 参数配置 ----
EXCEL_PATH = '/u01/liushiguo/lsg/Sonic/yoya生成内容评估-测试题集.xlsx'
SHEET_NAME = None  # 若为第一个工作表，保持 None，否则填写名称
OUTPUT_DIR = '/u01/liushiguo/lsg/Sonic/yoyadata'


def download_images():
    # 确保输出目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 加载工作簿和工作表
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    # 使用 openpyxl_image_loader 加载图片
    loader = SheetImageLoader(ws)

    # 遍历从第2行到最后一行
    for row in range(2, ws.max_row + 1):
        # A 列编号
        id_value = ws.cell(row=row, column=1).value
        # C 列单元格坐标
        cell = f'C{row}'

        # 检查 C 列是否有图片
        if loader.image_in(cell):
            img = loader.get(cell)
            # 构造输出文件名，统一保存为 PNG
            filename = f"{id_value}.png"
            save_path = os.path.join(OUTPUT_DIR, filename)
            # 保存图片
            img.save(save_path)
            print(f"已保存: {save_path}")
        else:
            print(f"第 {row} 行 C 列无图片，跳过。")

    print("所有图片下载完成。")


if __name__ == '__main__':
    download_images()
