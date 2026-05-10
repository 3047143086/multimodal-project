import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from tempfile import TemporaryDirectory
import cv2
from openai import OpenAI
import requests
from requests_toolbelt import multipart

# ---- 配置参数 ----
excel_path = "/u01/liushiguo/lsg/Sonic/tmp_unzip/yoya生成内容评估-测试题集.xlsx"
text_col = 4  # D列（文本）
image_col = 3  # C列（图片对象）
result_col = 5  # E列（结果）
temp_image_dir = "/u01/liushiguo/lsg/Sonic/temp_images"  # 临时图片存储目录

# ---- API配置 ----
bucket_key = "zs-a3efde7e"
api_key = "sk-73c002392bd94313b5c8c6da36ca1bc9"
base_url = "https://dashscope.aliyuncs.com/compatible-mode/v1"
upload_url = "http://hongqiplus.wengegroup.com/mam/api/file/getUrl"

def save_images_from_excel(worksheet):
    """从Excel提取并保存所有图片，返回行号到路径的映射字典"""
    image_map = {}
    
    # 创建临时目录
    os.makedirs(temp_image_dir, exist_ok=True)
    
    # 遍历所有图片对象
    for idx, img in enumerate(worksheet._images):
        # 获取图片锚定位置（openpyxl 3.0.10+）
        anchor = img.anchor
        if isinstance(anchor, openpyxl.drawing.spreadsheet_drawing.AnchorMarker):
            row = anchor.row + 1  # 转换为1-based行号
        else:  # 兼容旧版本处理
            row = img.anchor._from.row + 1
        
        # 生成唯一文件名
        filename = f"row_{row}_image_{idx}.png"
        save_path = os.path.join(temp_image_dir, filename)
        
        # 保存图片
        with open(save_path, "wb") as f:
            f.write(img._data())
        
        image_map[row] = save_path
    
    return image_map

def upload_data(img_path):
    """上传图片并获取URL"""
    wos_name = os.path.basename(img_path)
    try:
        data = multipart.MultipartEncoder(fields={
            'files': (wos_name, open(img_path, 'rb'), 'multipart/form-data')
        })
        headers = {
            "STORE_BUCKET_KEY": bucket_key,
            'Content-Type': data.content_type,
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1"
        }
        response = requests.post(upload_url, data=data, headers=headers)
        response.raise_for_status()
        return response.json()['data'][0]
    except Exception as e:
        print(f"上传失败：{str(e)}")
        return None

def process_excel():
    # 加载Excel文件
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 步骤1：提取并保存所有图片
    print("正在提取Excel中的图片...")
    image_map = save_images_from_excel(ws)
    print(f"成功保存{len(image_map)}张图片到临时目录")
    
    # 初始化OpenAI客户端
    client = OpenAI(api_key=api_key, base_url=base_url)
    
    # 步骤2：处理每一行数据
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # 获取当前行号（Excel中的实际行号）
        excel_row = row[0].row
        
        # 获取图片路径（从映射字典）
        image_path = image_map.get(excel_row)
        text_content = row[text_col-1].value  # D列值
        
        if not image_path or not text_content:
            continue
        
        # 上传图片获取URL
        file_url = upload_data(image_path)
        if not file_url:
            continue
        
        # 构建API请求
        try:
            completion = client.chat.completions.create(
                model="qwen-vl-plus-latest",
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": text_content},
                        {"type": "image_url", "image_url": {"url": file_url}}
                    ]
                }]
            )
            # 提取结果并写入E列
            result = completion.choices[0].message.content
            ws.cell(row=excel_row, column=result_col, value=result)
            print(f"成功处理第{excel_row}行")
        except Exception as e:
            print(f"第{excel_row}行处理失败：{str(e)}")
            ws.cell(row=excel_row, column=result_col, value=f"Error: {str(e)}")
    
    # 保存修改后的Excel文件（保留原始图片）
    wb.save(excel_path)
    print("处理完成，结果已保存！")
    
    # 可选：清理临时图片
    # shutil.rmtree(temp_image_dir)

if __name__ == "__main__":
    process_excel()